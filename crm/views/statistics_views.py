from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden, JsonResponse
from django.db.models import Count, Avg, F, ExpressionWrapper, fields, Q, Sum
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth, TruncYear
from django.utils import timezone
import datetime
from datetime import timedelta
import json
import logging
import csv
import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..models import (
    Ticket, ActivityLog, UserProfile, 
    Organization, TicketStatistics, AgentWorkLog, WorkHours
)
from ..views.error_views import forbidden_access

# Configure logger
logger = logging.getLogger(__name__)

@login_required
def statistics_dashboard(request):
    """View for displaying statistics dashboard"""
    user = request.user
    role = user.profile.role
    
    # Check if the user has access to statistics
    if role not in ['admin', 'superagent']:
        # Check if the user's group has explicit permission
        if not (user.groups.exists() and user.groups.first().settings.show_statistics):
            return forbidden_access(request, "statystyk")
    
    # Time period filter
    period = request.GET.get('period', 'month')  # Default to month
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Get current date for default periods
    today = timezone.now().date()
    
    # Set default date range based on period if not specified
    if not date_from or not date_to:
        if period == 'day':
            date_from = today
            date_to = today
        elif period == 'week':
            date_from = today - timedelta(days=today.weekday())
            date_to = date_from + timedelta(days=6)
        elif period == 'month':
            date_from = today.replace(day=1)
            next_month = today.month + 1 if today.month < 12 else 1
            next_month_year = today.year if today.month < 12 else today.year + 1
            date_to = datetime.date(next_month_year, next_month, 1) - timedelta(days=1)
        elif period == 'year':
            date_from = today.replace(month=1, day=1)
            date_to = today.replace(month=12, day=31)
    else:
        try:
            date_from = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
            date_to = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
        except ValueError:
            # If date parsing fails, use default for selected period
            if period == 'day':
                date_from = today
                date_to = today
            elif period == 'week':
                date_from = today - timedelta(days=today.weekday())
                date_to = date_from + timedelta(days=6)
            elif period == 'month':
                date_from = today.replace(day=1)
                next_month = today.month + 1 if today.month < 12 else 1
                next_month_year = today.year if today.month < 12 else today.year + 1
                date_to = datetime.date(next_month_year, next_month, 1) - timedelta(days=1)
            elif period == 'year':
                date_from = today.replace(month=1, day=1)
                date_to = today.replace(month=12, day=31)
    
    # Filter tickets based on user role and date range
    if role == 'admin':
        # Admin sees all tickets
        tickets = Ticket.objects.filter(
            created_at__date__gte=date_from,
            created_at__date__lte=date_to
        )
    elif role == 'superagent':
        # Filter for superagent - depends on what they have access to
        user_orgs = user.profile.organizations.all()
        tickets = Ticket.objects.filter(
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
            organization__in=user_orgs
        )
    else:
        # Fallback for other roles with explicit permission
        user_orgs = user.profile.organizations.all()
        tickets = Ticket.objects.filter(
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
            organization__in=user_orgs
        )
    
    # Organization filter
    org_filter = request.GET.get('organization', '')
    if org_filter:
        tickets = tickets.filter(organization_id=org_filter)
    
    # Agent filter
    agent_filter = request.GET.get('agent', '')
    if agent_filter:
        tickets = tickets.filter(assigned_to_id=agent_filter)
    
    # On Duty filter
    on_duty_filter = request.GET.get('on_duty', '')
    if on_duty_filter:
        if on_duty_filter == 'true':
            tickets = tickets.filter(on_duty=True)
        elif on_duty_filter == 'false':
            tickets = tickets.filter(on_duty=False)
    
    # Count tickets by status
    new_tickets = tickets.filter(status='new').count()
    in_progress_tickets = tickets.filter(status='in_progress').count()
    
    # Verify the filter value for unresolved tickets - first check if we need to use a different key
    # Some systems might use 'waiting' or 'reopened' as status names
    unresolved_count = tickets.filter(status='unresolved').count()
    
    # Debug logging for troubleshooting
    logger.info(f"Statistics: Found {unresolved_count} unresolved tickets")
    if unresolved_count == 0:
        # Check for tickets with similar statuses to see if we're missing something
        for status_check in ['waiting', 'reopened', 'problem', 'stuck']:
            alt_count = tickets.filter(status=status_check).count()
            if alt_count > 0:
                logger.info(f"Statistics: Found {alt_count} tickets with status '{status_check}'")

    unresolved_tickets = unresolved_count
    resolved_tickets = tickets.filter(status='resolved').count()
    closed_tickets = tickets.filter(status='closed').count()
    total_tickets = tickets.count()
    assigned_tickets = tickets.exclude(assigned_to__isnull=True).count()
    
    # Debug logging to verify totals
    logger.info(f"Statistics: Total tickets: {total_tickets}, Assigned tickets: {assigned_tickets}, Sum of statuses: {new_tickets + in_progress_tickets + unresolved_tickets + resolved_tickets + closed_tickets}")
    
    # Calculate ticket resolution metrics
    avg_resolution_time = tickets.exclude(
        resolved_at__isnull=True
    ).aggregate(
        avg_time=Avg(
            ExpressionWrapper(
                F('resolved_at') - F('created_at'),
                output_field=fields.DurationField()
            )
        )
    )['avg_time']
    
    if avg_resolution_time:
        avg_resolution_hours = avg_resolution_time.total_seconds() / 3600
    else:
        avg_resolution_hours = 0
    
    # Calculate average actual resolution time (rzeczywisty czas wykonania)
    avg_actual_resolution_time = tickets.exclude(
        actual_resolution_time__isnull=True
    ).aggregate(
        avg_actual_time=Avg('actual_resolution_time')
    )['avg_actual_time']
    
    if avg_actual_resolution_time:
        avg_actual_hours = float(avg_actual_resolution_time)
    else:
        avg_actual_hours = None
    
    # Count tickets with actual resolution time (for percentage calculation)
    tickets_with_actual_time = tickets.exclude(actual_resolution_time__isnull=True).count()
    tickets_with_actual_time_percentage = (tickets_with_actual_time / total_tickets * 100) if total_tickets > 0 else 0
    
    # Get priority distribution
    priority_distribution = tickets.values('priority').annotate(
        count=Count('id')
    ).order_by('priority')
    
    # Get category distribution
    category_distribution = tickets.values('category').annotate(
        count=Count('id')
    ).order_by('category')
    
    # Get tickets by creation date
    if period == 'day':
        tickets_by_date = tickets.annotate(
            date=TruncDay('created_at')
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date')
    elif period == 'week':
        tickets_by_date = tickets.annotate(
            date=TruncWeek('created_at')
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date')
    elif period == 'month':
        tickets_by_date = tickets.annotate(
            date=TruncMonth('created_at')
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date')
    else:  # year
        tickets_by_date = tickets.annotate(
            date=TruncYear('created_at')
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date')
    
    # Get agent performance metrics - for agents with assigned tickets
    agent_performance = []
    
    if role in ['admin', 'superagent']:
        agents_with_tickets = tickets.values(
            'assigned_to'
        ).exclude(
            assigned_to__isnull=True
        ).annotate(
            ticket_count=Count('id')
        ).order_by('-ticket_count')
        
        for agent_data in agents_with_tickets:
            agent_id = agent_data['assigned_to']
            if agent_id:
                try:
                    agent_user = UserProfile.objects.get(user_id=agent_id).user
                    agent_tickets = tickets.filter(assigned_to_id=agent_id)
                    
                    agent_resolved = agent_tickets.filter(status__in=['resolved', 'closed']).count()
                    agent_total = agent_tickets.count()
                    
                    if agent_total > 0:
                        resolution_rate = (agent_resolved / agent_total) * 100
                    else:
                        resolution_rate = 0
                    
                    # Calculate average resolution time for this agent
                    agent_avg_time = agent_tickets.exclude(
                        resolved_at__isnull=True
                    ).aggregate(
                        avg_time=Avg(
                            ExpressionWrapper(
                                F('resolved_at') - F('created_at'),
                                output_field=fields.DurationField()
                            )
                        )
                    )['avg_time']
                    
                    if agent_avg_time:
                        agent_avg_hours = agent_avg_time.total_seconds() / 3600
                    else:
                        agent_avg_hours = 0
                    
                    # Calculate average actual resolution time for this agent
                    agent_actual_avg_time = agent_tickets.exclude(
                        actual_resolution_time__isnull=True
                    ).aggregate(
                        avg_actual_time=Avg('actual_resolution_time')
                    )['avg_actual_time']
                    
                    if agent_actual_avg_time:
                        agent_actual_avg_hours = float(agent_actual_avg_time)
                    else:
                        agent_actual_avg_hours = None
                    
                    # Count tickets with actual time for this agent
                    agent_tickets_with_actual_time = agent_tickets.exclude(actual_resolution_time__isnull=True).count()
                    
                    agent_performance.append({
                        'agent_id': agent_id,  # Add this line
                        'agent_name': f"{agent_user.first_name} {agent_user.last_name}" if agent_user.first_name else agent_user.username,
                        'ticket_count': agent_total,
                        'resolved_count': agent_resolved,
                        'resolution_rate': resolution_rate,
                        'avg_resolution_time': agent_avg_hours,
                        'avg_actual_resolution_time': agent_actual_avg_hours,
                        'tickets_with_actual_time': agent_tickets_with_actual_time
                    })
                except UserProfile.DoesNotExist:
                    # Skip if agent profile doesn't exist
                    pass
    
    # Get organizations for filter dropdown
    organizations = []
    if role == 'admin':
        organizations = Organization.objects.all()
    else:
        organizations = user.profile.organizations.all()
    
    # Get agents for filter dropdown (either all agents for admin or agents in user's organizations)
    if role == 'admin':
        agents = UserProfile.objects.filter(role='agent').select_related('user')
    else:
        # For superagent, only show agents in their organizations
        user_orgs = user.profile.organizations.all()
        agents = UserProfile.objects.filter(
            role='agent',
            organizations__in=user_orgs
        ).distinct().select_related('user')
    
    # Calculate the estimated agent work time if we have agent work logs
    agent_work_time_stats = {}
    if AgentWorkLog.objects.exists():
        # Calculate average work time per ticket for each agent
        agent_logs = AgentWorkLog.objects.filter(
            ticket__in=tickets,
            start_time__date__gte=date_from,
            start_time__date__lte=date_to
        )
        
        agent_work_summary = agent_logs.values('agent').annotate(
            total_time=Sum('work_time_minutes'),
            ticket_count=Count('ticket', distinct=True)
        )
        
        for summary in agent_work_summary:
            agent_id = summary['agent']
            try:
                agent_user = UserProfile.objects.get(user_id=agent_id).user
                agent_name = f"{agent_user.first_name} {agent_user.last_name}" if agent_user.first_name else agent_user.username
                
                total_minutes = summary['total_time'] or 0
                ticket_count = summary['ticket_count']
                
                if ticket_count > 0:
                    avg_minutes_per_ticket = total_minutes / ticket_count
                else:
                    avg_minutes_per_ticket = 0
                
                agent_work_time_stats[agent_id] = {
                    'agent_name': agent_name,
                    'total_minutes': total_minutes,
                    'avg_minutes_per_ticket': avg_minutes_per_ticket,
                    'ticket_count': ticket_count
                }
            except UserProfile.DoesNotExist:
                pass
    
    # Format data for JSON serialization in charts
    priority_data = list(priority_distribution)
    category_data = list(category_distribution)
    
    # Format tickets_by_date for JSON serialization
    tickets_by_date_data = []
    for entry in tickets_by_date:
        tickets_by_date_data.append({
            'date': entry['date'].isoformat() if hasattr(entry['date'], 'isoformat') else str(entry['date']),
            'count': entry['count']
        })
    
    # Process agent performance data with work time stats
    for ap in agent_performance:
        agent_id = ap.get('agent_id')
        if agent_id and agent_id in agent_work_time_stats:
            ap['work_time_stats'] = agent_work_time_stats[agent_id]
        else:
            ap['work_time_stats'] = None
    
    # Calculate average tickets per agent (sum of all ticket counts / number of agents)
    avg_tickets_per_agent = 0
    if agent_performance:
        total_agent_tickets = sum(ap['ticket_count'] for ap in agent_performance)
        avg_tickets_per_agent = total_agent_tickets / len(agent_performance) if len(agent_performance) > 0 else 0
    
    context = {
        'period': period,
        'date_from': date_from,
        'date_to': date_to,
        'total_tickets': total_tickets,
        'assigned_tickets': assigned_tickets,
        'avg_tickets_per_agent': avg_tickets_per_agent,
        'new_tickets': new_tickets,
        'in_progress_tickets': in_progress_tickets,
        'unresolved_tickets': unresolved_tickets,  # Add this missing key
        'resolved_tickets': resolved_tickets,
        'closed_tickets': closed_tickets,
        'avg_resolution_hours': avg_resolution_hours,
        'avg_actual_hours': avg_actual_hours,
        'tickets_with_actual_time': tickets_with_actual_time,
        'tickets_with_actual_time_percentage': tickets_with_actual_time_percentage,
        'priority_distribution': priority_data,
        'category_distribution': category_data,
        'tickets_by_date': tickets_by_date_data,
        'agent_performance': agent_performance,
        'organizations': organizations,
        'agents': agents,
        'org_filter': org_filter,
        'agent_filter': agent_filter,
        'on_duty_filter': on_duty_filter,
        'agent_work_time_stats': agent_work_time_stats,
        'debug_unresolved_count': unresolved_count,  # Keep this for debugging
    }
    
    return render(request, 'crm/statistics/statistics_dashboard.html', context)

@login_required
def update_agent_work_log(request):
    """API endpoint to record agent work time on tickets"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    user = request.user
    if user.profile.role not in ['agent', 'admin', 'superagent']:
        return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)
    
    ticket_id = request.POST.get('ticket_id')
    action = request.POST.get('action')  # 'start' or 'stop'
    notes = request.POST.get('notes', '')
    
    try:
        ticket = Ticket.objects.get(id=ticket_id)
    except Ticket.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Ticket not found'}, status=404)
    
    # Only assigned agent, admin or superagent can log work time
    if user.profile.role not in ['admin', 'superagent'] and ticket.assigned_to != user:
        return JsonResponse({'status': 'error', 'message': 'You are not assigned to this ticket'}, status=403)
    
    now = timezone.now()
    
    if action == 'start':
        # Check if there's already an open work log
        open_log = AgentWorkLog.objects.filter(
            agent=user,
            ticket=ticket,
            end_time__isnull=True
        ).first()
        
        if open_log:
            return JsonResponse({
                'status': 'error',
                'message': 'You already have an open work log for this ticket'
            }, status=400)
        
        # Start new work log
        work_log = AgentWorkLog.objects.create(
            agent=user,
            ticket=ticket,
            start_time=now,
            notes=notes
        )
        
        return JsonResponse({
            'status': 'success',
            'message': 'Work log started',
            'work_log_id': work_log.id
        })
        
    elif action == 'stop':
        # Find open work log
        open_log = AgentWorkLog.objects.filter(
            agent=user,
            ticket=ticket,
            end_time__isnull=True
        ).first()
        
        if not open_log:
            return JsonResponse({
                'status': 'error',
                'message': 'No open work log found for this ticket'
            }, status=400)
        
        # Update end time and calculate work time
        open_log.end_time = now
        
        # Calculate work time considering only work hours
        work_minutes = calculate_work_minutes(open_log.start_time, now)
        open_log.work_time_minutes = work_minutes
        
        if notes:
            open_log.notes = notes
            
        open_log.save()
        
        return JsonResponse({
            'status': 'success',
            'message': 'Work log completed',
            'work_minutes': work_minutes
        })
    
    return JsonResponse({'status': 'error', 'message': 'Invalid action'}, status=400)

def calculate_work_minutes(start_time, end_time):
    """Calculate working minutes between two timestamps, considering only work hours"""
    # Get work hours configuration
    work_hours = WorkHours.objects.all()
    
    # If no work hours defined, use 8:00-16:00 Mon-Fri as default
    if not work_hours.exists():
        default_work_hours = {}
        for day in range(5):  # Monday to Friday
            default_work_hours[day] = [(datetime.time(8, 0), datetime.time(16, 0))]
    else:
        # Organize work hours by day
        default_work_hours = {}
        for wh in work_hours:
            if wh.is_working_day:
                if wh.day_of_week not in default_work_hours:
                    default_work_hours[wh.day_of_week] = []
                default_work_hours[wh.day_of_week].append((wh.start_time, wh.end_time))
    
    # If start and end are on the same day, calculate directly
    if start_time.date() == end_time.date():
        day_of_week = start_time.weekday()
        if day_of_week in default_work_hours:
            # Check each work period in this day
            total_minutes = 0
            for start_hour, end_hour in default_work_hours[day_of_week]:
                period_start = datetime.datetime.combine(
                    start_time.date(), 
                    start_hour,
                    tzinfo=start_time.tzinfo
                )
                period_end = datetime.datetime.combine(
                    start_time.date(), 
                    end_hour,
                    tzinfo=start_time.tzinfo
                )
                
                # Find overlap between work period and log period
                overlap_start = max(period_start, start_time)
                overlap_end = min(period_end, end_time)
                
                if overlap_end > overlap_start:
                    overlap_minutes = (overlap_end - overlap_start).total_seconds() / 60
                    total_minutes += overlap_minutes
            
            return total_minutes
        else:
            # Not a working day
            return 0
    else:
        # If spanning multiple days, calculate day by day
        current_date = start_time.date()
        end_date = end_time.date()
        total_minutes = 0
        
        while current_date <= end_date:
            if current_date == start_time.date():
                # First day - from start_time to end of work day
                day_start = start_time
                day_end = datetime.datetime.combine(
                    current_date,
                    datetime.time(23, 59, 59),
                    tzinfo=start_time.tzinfo
                )
            elif current_date == end_date:
                # Last day - from start of work day to end_time
                day_start = datetime.datetime.combine(
                    current_date,
                    datetime.time(0, 0),
                    tzinfo=start_time.tzinfo
                )
                day_end = end_time
            else:
                # Middle days - full day
                day_start = datetime.datetime.combine(
                    current_date,
                    datetime.time(0, 0),
                    tzinfo=start_time.tzinfo
                )
                day_end = datetime.datetime.combine(
                    current_date,
                    datetime.time(23, 59, 59),
                    tzinfo=start_time.tzinfo
                )
            
            # Calculate minutes for this day
            day_of_week = current_date.weekday()
            if day_of_week in default_work_hours:
                for start_hour, end_hour in default_work_hours[day_of_week]:
                    period_start = datetime.datetime.combine(
                        current_date, 
                        start_hour,
                        tzinfo=start_time.tzinfo
                    )
                    period_end = datetime.datetime.combine(
                        current_date, 
                        end_hour,
                        tzinfo=start_time.tzinfo
                    )
                    
                    overlap_start = max(period_start, day_start)
                    overlap_end = min(period_end, day_end)
                    
                    if overlap_end > overlap_start:
                        overlap_minutes = (overlap_end - overlap_start).total_seconds() / 60
                        total_minutes += overlap_minutes
            
            current_date += timedelta(days=1)
        
        return total_minutes

@login_required
def generate_statistics_report(request):
    """Generate and download statistics report"""
    user = request.user
    
    logger.info(f"Report generation started by user: {user.username} (role: {user.profile.role})")
    
    # Only admin and superagent can generate reports
    if user.profile.role not in ['admin', 'superagent']:
        logger.warning(f"Permission denied for user {user.username} with role {user.profile.role}")
        return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)
    
    try:
        period_type = request.POST.get('period_type', 'month')
        period_start = request.POST.get('period_start')
        period_end = request.POST.get('period_end')
        organization_id = request.POST.get('organization')
        agent_id = request.POST.get('agent')
        on_duty_filter = request.POST.get('on_duty', '')  # on_duty filter
        report_format = request.POST.get('format', 'xlsx')  # xlsx or csv
        
        logger.info(f"Report parameters: period_type={period_type}, period_start={period_start}, period_end={period_end}, organization_id={organization_id}, agent_id={agent_id}, on_duty={on_duty_filter}, format={report_format}")
        
        # Validate input parameters
        if not period_start or not period_end:
            logger.error("Missing required date parameters")
            return JsonResponse({'status': 'error', 'message': 'Brakuje wymaganych dat'}, status=400)
        
        try:
            period_start_date = datetime.datetime.strptime(period_start, '%Y-%m-%d').date()
            period_end_date = datetime.datetime.strptime(period_end, '%Y-%m-%d').date()
            logger.info(f"Parsed dates: {period_start_date} to {period_end_date}")
        except (ValueError, TypeError) as e:
            logger.error(f"Date parsing error: {e}")
            return JsonResponse({'status': 'error', 'message': 'Nieprawidłowy format daty'}, status=400)
        
        # Check date range validity
        if period_start_date > period_end_date:
            logger.error("Start date is after end date")
            return JsonResponse({'status': 'error', 'message': 'Data początkowa nie może być późniejsza niż końcowa'}, status=400)
        
        # Filter tickets
        logger.info("Building tickets query...")
        tickets_query = Ticket.objects.filter(
            created_at__date__gte=period_start_date,
            created_at__date__lte=period_end_date
        )
        
        initial_count = tickets_query.count()
        logger.info(f"Initial tickets count for date range: {initial_count}")
        
        # Apply organization filter if specified
        organization = None
        if organization_id:
            try:
                organization = Organization.objects.get(id=organization_id)
                tickets_query = tickets_query.filter(organization=organization)
                org_filtered_count = tickets_query.count()
                logger.info(f"After organization filter ({organization.name}): {org_filtered_count} tickets")
            except Organization.DoesNotExist:
                logger.error(f"Organization not found: {organization_id}")
                return JsonResponse({'status': 'error', 'message': 'Organizacja nie została znaleziona'}, status=400)
            except Exception as e:
                logger.error(f"Error filtering by organization: {e}")
                return JsonResponse({'status': 'error', 'message': f'Błąd filtrowania organizacji: {str(e)}'}, status=400)
        
        # Apply agent filter if specified
        agent = None
        if agent_id:
            try:
                agent = UserProfile.objects.get(user_id=agent_id).user
                tickets_query = tickets_query.filter(assigned_to=agent)
                agent_filtered_count = tickets_query.count()
                logger.info(f"After agent filter ({agent.username}): {agent_filtered_count} tickets")
            except UserProfile.DoesNotExist:
                logger.error(f"Agent not found: {agent_id}")
                return JsonResponse({'status': 'error', 'message': 'Agent nie został znaleziony'}, status=400)
            except Exception as e:
                logger.error(f"Error filtering by agent: {e}")
                return JsonResponse({'status': 'error', 'message': f'Błąd filtrowania agenta: {str(e)}'}, status=400)
        
        # Apply on_duty filter if specified
        if on_duty_filter:
            if on_duty_filter == 'true':
                tickets_query = tickets_query.filter(on_duty=True)
                logger.info(f"After on_duty=True filter: {tickets_query.count()} tickets")
            elif on_duty_filter == 'false':
                tickets_query = tickets_query.filter(on_duty=False)
                logger.info(f"After on_duty=False filter: {tickets_query.count()} tickets")
        
        # Calculate statistics
        logger.info("Calculating ticket statistics...")
        try:
            tickets_opened = tickets_query.count()
            tickets_closed = tickets_query.filter(status='closed').count()
            tickets_resolved = tickets_query.filter(status='resolved').count()
            tickets_new = tickets_query.filter(status='new').count()
            tickets_in_progress = tickets_query.filter(status='in_progress').count()
            tickets_unresolved = tickets_query.filter(status='unresolved').count()
            
            logger.info(f"Ticket counts: total={tickets_opened}, new={tickets_new}, in_progress={tickets_in_progress}, unresolved={tickets_unresolved}, resolved={tickets_resolved}, closed={tickets_closed}")
            
            # Verify total
            sum_statuses = tickets_new + tickets_in_progress + tickets_unresolved + tickets_resolved + tickets_closed
            if sum_statuses != tickets_opened:
                logger.warning(f"Status counts don't match total: {sum_statuses} vs {tickets_opened}")
                # Check for other statuses
                all_statuses = tickets_query.values('status').annotate(count=Count('id'))
                logger.info(f"All status counts: {list(all_statuses)}")
        except Exception as e:
            logger.error(f"Error calculating basic statistics: {e}")
            return JsonResponse({'status': 'error', 'message': f'Błąd obliczania statystyk: {str(e)}'}, status=500)
        
        # Calculate average resolution time
        logger.info("Calculating average resolution time...")
        try:
            resolution_time_data = tickets_query.exclude(
                resolved_at__isnull=True
            ).aggregate(
                avg_time=Avg(
                    ExpressionWrapper(
                        F('resolved_at') - F('created_at'),
                        output_field=fields.DurationField()
                    )
                )
            )
            avg_resolution_time = 0
            if resolution_time_data['avg_time']:
                avg_resolution_time = resolution_time_data['avg_time'].total_seconds() / 3600  # Convert to hours
            
            logger.info(f"Average resolution time: {avg_resolution_time} hours")
        except Exception as e:
            logger.error(f"Error calculating resolution time: {e}")
            avg_resolution_time = 0
        
        # Calculate priority distribution
        logger.info("Calculating priority distribution...")
        try:
            priority_counts = tickets_query.values('priority').annotate(count=Count('id'))
            priority_distribution = {item['priority']: item['count'] for item in priority_counts}
            logger.info(f"Priority distribution: {priority_distribution}")
        except Exception as e:
            logger.error(f"Error calculating priority distribution: {e}")
            priority_distribution = {}
        
        # Calculate category distribution
        logger.info("Calculating category distribution...")
        try:
            category_counts = tickets_query.values('category').annotate(count=Count('id'))
            category_distribution = {item['category']: item['count'] for item in category_counts}
            logger.info(f"Category distribution: {category_distribution}")
        except Exception as e:
            logger.error(f"Error calculating category distribution: {e}")
            category_distribution = {}
        
        # Calculate average agent work time if agent work logs exist
        logger.info("Calculating agent work time...")
        try:
            avg_agent_work_time = 0
            if AgentWorkLog.objects.filter(ticket__in=tickets_query).exists():
                agent_work_data = AgentWorkLog.objects.filter(
                    ticket__in=tickets_query
                ).aggregate(
                    avg_time=Avg('work_time_minutes')
                )
                avg_agent_work_time = agent_work_data['avg_time'] or 0
            logger.info(f"Average agent work time: {avg_agent_work_time} minutes")
        except Exception as e:
            logger.error(f"Error calculating agent work time: {e}")
            avg_agent_work_time = 0
        
        # Get agent performance data
        logger.info("Calculating agent performance...")
        agent_performance = []
        try:
            if user.profile.role in ['admin', 'superagent']:
                agents_with_tickets = tickets_query.values(
                    'assigned_to'
                ).exclude(
                    assigned_to__isnull=True
                ).annotate(
                    ticket_count=Count('id')
                ).order_by('-ticket_count')
                
                logger.info(f"Found {len(agents_with_tickets)} agents with tickets")
                
                for agent_data in agents_with_tickets:
                    agent_id = agent_data['assigned_to']
                    if agent_id:
                        try:
                            agent_user = UserProfile.objects.get(user_id=agent_id).user
                            agent_tickets = tickets_query.filter(assigned_to_id=agent_id)
                            
                            agent_resolved = agent_tickets.filter(status__in=['resolved', 'closed']).count()
                            agent_total = agent_tickets.count()
                            
                            if agent_total > 0:
                                resolution_rate = (agent_resolved / agent_total) * 100
                            else:
                                resolution_rate = 0
                            

                            # Calculate average resolution time for this agent
                            agent_avg_time = agent_tickets.exclude(
                                resolved_at__isnull=True
                            ).aggregate(
                                avg_time=Avg(
                                    ExpressionWrapper(
                                        F('resolved_at') - F('created_at'),
                                        output_field=fields.DurationField()
                                    )
                                )
                            )['avg_time']
                            
                            if agent_avg_time:
                                agent_avg_hours = agent_avg_time.total_seconds() / 3600
                            else:
                                agent_avg_hours = 0
                            
                            # Calculate average actual resolution time for this agent
                            agent_actual_avg_time = agent_tickets.exclude(
                                actual_resolution_time__isnull=True
                            ).aggregate(
                                avg_actual_time=Avg('actual_resolution_time')
                            )['avg_actual_time']
                            
                            if agent_actual_avg_time:
                                agent_actual_avg_hours = float(agent_actual_avg_time)
                            else:
                                agent_actual_avg_hours = None
                            
                            # Count tickets with actual time for this agent
                            agent_tickets_with_actual_time = agent_tickets.exclude(actual_resolution_time__isnull=True).count()

                            agent_performance.append({
                                'agent_id': agent_id,  # Add agent_id for ticket filtering
                                'agent_name': f"{agent_user.first_name} {agent_user.last_name}" if agent_user.first_name else agent_user.username,
                                'ticket_count': agent_total,
                                'resolved_count': agent_resolved,
                                'resolution_rate': resolution_rate,
                                'avg_resolution_time': agent_avg_hours,
                                'avg_actual_resolution_time': agent_actual_avg_hours,
                                'tickets_with_actual_time': agent_tickets_with_actual_time
                            })
                            
                            logger.debug(f"Agent {agent_user.username} (id={agent_id}): {agent_total} tickets, {resolution_rate}% resolution rate")
                        except UserProfile.DoesNotExist:
                            logger.warning(f"UserProfile not found for user_id: {agent_id}")
                        except Exception as e:
                            logger.error(f"Error processing agent {agent_id}: {e}")
                
                logger.info(f"Processed {len(agent_performance)} agents for performance data")
        except Exception as e:
            logger.error(f"Error calculating agent performance: {e}")
            agent_performance = []
        
        # Create the report file
        logger.info(f"Generating {report_format} report...")
        try:
            if report_format == 'csv':
                return _generate_csv_report(
                    period_start_date, period_end_date, organization, agent,
                    tickets_opened, tickets_closed, tickets_resolved, tickets_new, 
                    tickets_in_progress, tickets_unresolved, avg_resolution_time,
                    priority_distribution, category_distribution, agent_performance
                )
            else:  # xlsx
                return _generate_excel_report(
                    period_start_date, period_end_date, organization, agent,
                    tickets_opened, tickets_closed, tickets_resolved, tickets_new,
                    tickets_in_progress, tickets_unresolved, avg_resolution_time,
                    priority_distribution, category_distribution, agent_performance
                )
        except ImportError as e:
            logger.error(f"Import error during report generation: {e}")
            return JsonResponse({
                'status': 'error', 
                'message': f'Brak wymaganej biblioteki: {str(e)}. Spróbuj użyć formatu CSV.'
            }, status=500)
        except Exception as e:
            logger.error(f"Error generating {report_format} report: {e}")
            import traceback
            logger.error(f"Full traceback: {traceback.format_exc()}")
            return JsonResponse({
                'status': 'error', 
                'message': f'Błąd generowania raportu {report_format.upper()}: {str(e)}'
            }, status=500)
            
    except Exception as e:
        logger.error(f"Unexpected error in generate_statistics_report: {e}")
        import traceback
        logger.error(f"Full traceback: {traceback.format_exc()}")
        return JsonResponse({
            'status': 'error', 
            'message': f'Nieoczekiwany błąd: {str(e)}'
        }, status=500)

def _generate_csv_report(period_start, period_end, organization, agent, 
                        tickets_opened, tickets_closed, tickets_resolved, tickets_new,
                        tickets_in_progress, tickets_unresolved, avg_resolution_time,
                        priority_distribution, category_distribution, agent_performance):
    """Generate CSV report"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    
    # Generate filename
    org_name = organization.name if organization else "Wszystkie"
    agent_name = agent.username if agent else "Wszyscy"
    filename = f"raport_statystyk_{period_start}_{period_end}_{org_name}_{agent_name}.csv"
    filename = filename.replace(" ", "_").replace("/", "_")
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Header information
    writer.writerow(['RAPORT STATYSTYK ZGŁOSZEŃ'])
    writer.writerow([''])
    writer.writerow(['Okres:', f"{period_start} - {period_end}"])
    writer.writerow(['Organizacja:', org_name])
    writer.writerow(['Agent:', agent_name])
    writer.writerow(['Data generacji:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow([''])
    
    # Summary statistics
    writer.writerow(['PODSUMOWANIE'])
    writer.writerow(['Łącznie zgłoszeń:', tickets_opened])
    writer.writerow(['Nowych:', tickets_new])
    writer.writerow(['W trakcie:', tickets_in_progress])
    writer.writerow(['Nierozwiązanych:', tickets_unresolved])
    writer.writerow(['Rozwiązanych:', tickets_resolved])
    writer.writerow(['Zamkniętych:', tickets_closed])
    writer.writerow(['Średni czas rozwiązania (godziny):', f"{avg_resolution_time:.2f}"])
    writer.writerow([''])
    
    # Priority distribution
    writer.writerow(['ROZKŁAD WEDŁUG PRIORYTETU'])
    writer.writerow(['Priorytet', 'Liczba zgłoszeń'])
    priority_labels = {'low': 'Niski', 'medium': 'Średni', 'high': 'Wysoki', 'critical': 'Krytyczny'}
    for priority, count in priority_distribution.items():
        writer.writerow([priority_labels.get(priority, priority), count])
    writer.writerow([''])
    
    # Category distribution
    writer.writerow(['ROZKŁAD WEDŁUG KATEGORII'])
    writer.writerow(['Kategoria', 'Liczba zgłoszeń'])
    category_labels = {'hardware': 'Sprzęt', 'software': 'Oprogramowanie', 'network': 'Sieć', 'account': 'Konto', 'other': 'Inne'}
    for category, count in category_distribution.items():
        writer.writerow([category_labels.get(category, category), count])
    writer.writerow([''])
    
    # Agent performance with tickets details
    if agent_performance:
        writer.writerow(['WYDAJNOŚĆ AGENTÓW'])
        writer.writerow(['Agent', 'Liczba zgłoszeń', 'Rozwiązanych', '% rozwiązanych', 'Śr. czas rozwiązania (godz.)', 'Śr. rzeczywisty czas (godz.)', 'Zgł. z rzecz. czasem'])
        for ap in agent_performance:
            avg_actual = f"{ap['avg_actual_resolution_time']:.2f}" if ap.get('avg_actual_resolution_time') else "Brak danych"
            tickets_actual = ap.get('tickets_with_actual_time', 0)
            
            writer.writerow([
                ap['agent_name'], 
                ap['ticket_count'], 
                ap['resolved_count'],
                f"{ap['resolution_rate']:.1f}%",
                f"{ap['avg_resolution_time']:.2f}",
                avg_actual,
                tickets_actual
            ])
            
            # Add agent's tickets details
            writer.writerow([''])
            writer.writerow([f"  Zgłoszenia agenta: {ap['agent_name']}"])
            writer.writerow(['  ID', 'Tytuł', 'Status', 'Priorytet', 'Kategoria', 'Dyżur', 'Utworzono', 'Rozwiązano', 'Zamknięto', 'Rzecz. czas (h)'])
            
            # Get agent's tickets in period
            agent_id = ap.get('agent_id')
            if agent_id:
                # period_start and period_end are already date objects, use them directly
                agent_tickets = Ticket.objects.filter(
                    assigned_to_id=agent_id,
                    created_at__date__gte=period_start,
                    created_at__date__lte=period_end
                ).order_by('-created_at')
                
                if agent_tickets.exists():
                    status_labels = {'new': 'Nowy', 'in_progress': 'W trakcie', 'unresolved': 'Nierozwiązany', 'resolved': 'Rozwiązany', 'closed': 'Zamknięty'}
                    priority_labels = {'low': 'Niski', 'medium': 'Średni', 'high': 'Wysoki', 'critical': 'Krytyczny'}
                    category_labels = {'hardware': 'Sprzęt', 'software': 'Oprogramowanie', 'network': 'Sieć', 'account': 'Konto', 'other': 'Inne'}
                    
                    for ticket in agent_tickets:
                        writer.writerow([
                            f'  #{ticket.id}',
                            ticket.title[:50] + ('...' if len(ticket.title) > 50 else ''),
                            status_labels.get(ticket.status, ticket.status),
                            priority_labels.get(ticket.priority, ticket.priority),
                            category_labels.get(ticket.category, ticket.category),
                            'Tak' if ticket.on_duty else 'Nie',
                            ticket.created_at.strftime('%Y-%m-%d %H:%M'),
                            ticket.resolved_at.strftime('%Y-%m-%d %H:%M') if ticket.resolved_at else '-',
                            ticket.closed_at.strftime('%Y-%m-%d %H:%M') if ticket.closed_at else '-',
                            f"{ticket.actual_resolution_time:.2f}" if ticket.actual_resolution_time else '-'
                        ])
                else:
                    writer.writerow(['  ', 'Brak zgłoszeń w wybranym okresie'])
            
            writer.writerow([''])
    
    return response

def _generate_excel_report(period_start, period_end, organization, agent,
                          tickets_opened, tickets_closed, tickets_resolved, tickets_new,
                          tickets_in_progress, tickets_unresolved, avg_resolution_time,
                          priority_distribution, category_distribution, agent_performance):
    """Generate Excel report with formatting"""
    logger.info("Starting Excel report generation...")
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Raport Statystyk"
        logger.info("Created Excel workbook")
        
        # Styles
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=12)
        bold_font = Font(bold=True)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Generate filename
        org_name = organization.name if organization else "Wszystkie"
        agent_name = agent.username if agent else "Wszyscy"
        
        # Clean filename for Windows compatibility
        org_name = org_name.replace(" ", "_").replace("/", "_").replace("\\", "_").replace(":", "_")
        agent_name = agent_name.replace(" ", "_").replace("/", "_").replace("\\", "_").replace(":", "_")
        
        logger.info(f"Report for organization: {org_name}, agent: {agent_name}")
        
        # Header information
        ws['A1'] = 'RAPORT STATYSTYK ZGŁOSZEŃ'
        ws['A1'].font = header_font
        
        row = 3
        ws[f'A{row}'] = 'Okres:'
        ws[f'B{row}'] = f"{period_start} - {period_end}"
        row += 1
        ws[f'A{row}'] = 'Organizacja:'
        ws[f'B{row}'] = organization.name if organization else "Wszystkie"
        row += 1
        ws[f'A{row}'] = 'Agent:'
        ws[f'B{row}'] = agent.username if agent else "Wszyscy"
        row += 1
        ws[f'A{row}'] = 'Data generacji:'
        ws[f'B{row}'] = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        
        row += 3
        
        logger.info("Adding summary statistics to Excel...")
        
        # Summary statistics
        ws[f'A{row}'] = 'PODSUMOWANIE'
        ws[f'A{row}'].font = subheader_font
        row += 1
        
        summary_data = [
            ('Łącznie zgłoszeń:', tickets_opened),
            ('Nowych:', tickets_new),
            ('W trakcie:', tickets_in_progress),
            ('Nierozwiązanych:', tickets_unresolved),
            ('Rozwiązanych:', tickets_resolved),
            ('Zamkniętych:', tickets_closed),
            ('Średni czas rozwiązania (godziny):', f"{avg_resolution_time:.2f}"),
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = bold_font
            row += 1
        
        row += 2
        
        logger.info("Adding priority distribution to Excel...")
        
        # Priority distribution
        ws[f'A{row}'] = 'ROZKŁAD WEDŁUG PRIORYTETU'
        ws[f'A{row}'].font = subheader_font
        row += 1
        
        ws[f'A{row}'] = 'Priorytet'
        ws[f'B{row}'] = 'Liczba zgłoszeń'
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'].font = bold_font
        row += 1
        
        priority_labels = {'low': 'Niski', 'medium': 'Średni', 'high': 'Wysoki', 'critical': 'Krytyczny'}
        for priority, count in priority_distribution.items():
            ws[f'A{row}'] = priority_labels.get(priority, priority)
            ws[f'B{row}'] = count
            row += 1
        
        row += 2
        
        logger.info("Adding category distribution to Excel...")
        
        # Category distribution
        ws[f'A{row}'] = 'ROZKŁAD WEDŁUG KATEGORII'
        ws[f'A{row}'].font = subheader_font
        row += 1
        
        ws[f'A{row}'] = 'Kategoria'
        ws[f'B{row}'] = 'Liczba zgłoszeń'
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'].font = bold_font
        row += 1
        
        category_labels = {'hardware': 'Sprzęt', 'software': 'Oprogramowanie', 'network': 'Sieć', 'account': 'Konto', 'other': 'Inne'}
        for category, count in category_distribution.items():
            ws[f'A{row}'] = category_labels.get(category, category)
            ws[f'B{row}'] = count
            row += 1
        
        # Agent performance with tickets details
        if agent_performance:
            row += 2
            logger.info(f"Adding agent performance data ({len(agent_performance)} agents) to Excel...")
            
            ws[f'A{row}'] = 'WYDAJNOŚĆ AGENTÓW'
            ws[f'A{row}'].font = subheader_font
            row += 1
            
            headers = ['Agent', 'Liczba zgłoszeń', 'Rozwiązanych', '% rozwiązanych', 'Śr. czas rozwiązania (godz.)', 'Śr. rzeczywisty czas (godz.)', 'Zgł. z rzecz. czasem']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = bold_font
            row += 1
            
            # Colors for status badges (using global PatternFill import)
            status_colors = {
                'new': PatternFill(start_color="007bff", end_color="007bff", fill_type="solid"),
                'in_progress': PatternFill(start_color="17a2b8", end_color="17a2b8", fill_type="solid"),
                'unresolved': PatternFill(start_color="ffc107", end_color="ffc107", fill_type="solid"),
                'resolved': PatternFill(start_color="28a745", end_color="28a745", fill_type="solid"),
                'closed': PatternFill(start_color="6c757d", end_color="6c757d", fill_type="solid")
            }
            
            priority_colors = {
                'low': PatternFill(start_color="6c757d", end_color="6c757d", fill_type="solid"),
                'medium': PatternFill(start_color="17a2b8", end_color="17a2b8", fill_type="solid"),
                'high': PatternFill(start_color="ffc107", end_color="ffc107", fill_type="solid"),
                'critical': PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
            }
            
            white_font = Font(color="FFFFFF", bold=True)
            
            logger.info(f"Excel Report - Processing {len(agent_performance)} agents")
            
            for ap in agent_performance:
                logger.info(f"Agent {ap['agent_name']}: avg_actual_resolution_time = {ap.get('avg_actual_resolution_time')}, tickets_with_actual_time = {ap.get('tickets_with_actual_time', 0)}")
                avg_actual = f"{ap['avg_actual_resolution_time']:.2f}" if ap.get('avg_actual_resolution_time') else "Brak danych"
                tickets_actual = ap.get('tickets_with_actual_time', 0)
                
                ws[f'A{row}'] = ap['agent_name']
                ws[f'A{row}'].font = bold_font
                ws[f'B{row}'] = ap['ticket_count']
                ws[f'C{row}'] = ap['resolved_count']
                ws[f'D{row}'] = f"{ap['resolution_rate']:.1f}%"
                ws[f'E{row}'] = f"{ap['avg_resolution_time']:.2f}"
                ws[f'F{row}'] = avg_actual
                ws[f'G{row}'] = tickets_actual
                row += 1
                
                # Add agent's tickets details
                row += 1
                ws[f'A{row}'] = f"Zgłoszenia agenta: {ap['agent_name']}"
                ws[f'A{row}'].font = Font(bold=True, italic=True)
                row += 1
                
                ticket_headers = ['ID', 'Tytuł', 'Status', 'Priorytet', 'Kategoria', 'Dyżur', 'Utworzono', 'Rozwiązano', 'Zamknięto', 'Rzecz. czas (h)']
                for col, header in enumerate(ticket_headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = bold_font
                    cell.fill = PatternFill(start_color="e9ecef", end_color="e9ecef", fill_type="solid")
                row += 1
                
                # Get agent's tickets in period
                agent_id = ap.get('agent_id')
                logger.info(f"Excel Report - Processing agent: {ap.get('agent_name')}, agent_id={agent_id}")
                logger.info(f"Excel Report - Period: {period_start} to {period_end} (types: {type(period_start)}, {type(period_end)})")
                
                if agent_id:
                    # period_start and period_end are already date objects, use them directly
                    agent_tickets = Ticket.objects.filter(
                        assigned_to_id=agent_id,
                        created_at__date__gte=period_start,
                        created_at__date__lte=period_end
                    ).order_by('-created_at')
                    
                    logger.info(f"Excel Report - Found {agent_tickets.count()} tickets for agent {agent_id}")
                    
                    if agent_tickets.exists():
                        logger.info(f"Excel Report - Writing {agent_tickets.count()} tickets to Excel")
                        status_labels = {'new': 'Nowy', 'in_progress': 'W trakcie', 'unresolved': 'Nierozwiązany', 'resolved': 'Rozwiązany', 'closed': 'Zamknięty'}
                        priority_labels = {'low': 'Niski', 'medium': 'Średni', 'high': 'Wysoki', 'critical': 'Krytyczny'}
                        category_labels = {'hardware': 'Sprzęt', 'software': 'Oprogramowanie', 'network': 'Sieć', 'account': 'Konto', 'other': 'Inne'}
                        
                        for ticket in agent_tickets:
                            ws[f'A{row}'] = f'#{ticket.id}'
                            ws[f'B{row}'] = ticket.title[:50] + ('...' if len(ticket.title) > 50 else '')
                            
                            # Status with color
                            status_cell = ws[f'C{row}']
                            status_cell.value = status_labels.get(ticket.status, ticket.status)
                            if ticket.status in status_colors:
                                status_cell.fill = status_colors[ticket.status]
                                status_cell.font = white_font
                            
                            # Priority with color
                            priority_cell = ws[f'D{row}']
                            priority_cell.value = priority_labels.get(ticket.priority, ticket.priority)
                            if ticket.priority in priority_colors:
                                priority_cell.fill = priority_colors[ticket.priority]
                                priority_cell.font = white_font
                            
                            ws[f'E{row}'] = category_labels.get(ticket.category, ticket.category)
                            
                            # On Duty with color
                            on_duty_cell = ws[f'F{row}']
                            on_duty_cell.value = 'Tak' if ticket.on_duty else 'Nie'
                            if ticket.on_duty:
                                on_duty_cell.fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
                                on_duty_cell.font = white_font
                            else:
                                on_duty_cell.fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
                                on_duty_cell.font = white_font
                            
                            ws[f'G{row}'] = ticket.created_at.strftime('%Y-%m-%d %H:%M')
                            ws[f'H{row}'] = ticket.resolved_at.strftime('%Y-%m-%d %H:%M') if ticket.resolved_at else '-'
                            ws[f'I{row}'] = ticket.closed_at.strftime('%Y-%m-%d %H:%M') if ticket.closed_at else '-'
                            ws[f'J{row}'] = f"{ticket.actual_resolution_time:.2f}" if ticket.actual_resolution_time else '-'
                            row += 1
                    else:
                        logger.warning(f"Excel Report - No tickets found for agent {agent_id} in period")
                        ws[f'A{row}'] = 'Brak zgłoszeń w wybranym okresie'
                        ws[f'A{row}'].font = Font(italic=True, color="6c757d")
                        row += 1
                else:
                    logger.error(f"Excel Report - agent_id is missing! Agent data: {ap}")
                    ws[f'A{row}'] = 'Błąd: Brak ID agenta'
                    ws[f'A{row}'].font = Font(italic=True, color="dc3545")
                    row += 1
                
                row += 1
        
        logger.info("Auto-adjusting column widths...")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Preparing HTTP response...")
        
        # Create response with proper headers
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Generate clean filename
        filename = f"raport_statystyk_{period_start}_{period_end}_{org_name}_{agent_name}.xlsx"
        
        # Set proper headers for file download
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Description'] = 'File Transfer'
        response['Content-Transfer-Encoding'] = 'binary'
        response['Cache-Control'] = 'must-revalidate, post-check=0, pre-check=0'
        response['Pragma'] = 'public'
        
        logger.info(f"Saving Excel file with filename: {filename}")
        
        # Save workbook to response
        try:
            wb.save(response)
            logger.info("Excel file saved to response successfully")
        except Exception as save_error:
            logger.error(f"Error saving workbook to response: {save_error}")
            raise
        
        # Log response details
        logger.info(f"Response content type: {response.get('Content-Type')}")
        logger.info(f"Response content disposition: {response.get('Content-Disposition')}")
        
        logger.info("Excel report generation completed successfully")
        return response
        
    except Exception as e:
        logger.error(f"Error in _generate_excel_report: {e}")
        import traceback
        logger.error(f"Full traceback: {traceback.format_exc()}")
        raise


@login_required
def generate_organization_report(request):
    """
    Generate Excel report showing actual resolution time by organization
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Metoda POST wymagana'}, status=405)
    
    try:
        user = request.user
        role = user.profile.role
        
        # Only admins and superagents can generate this report
        if role not in ['admin', 'superagent']:
            return JsonResponse({'status': 'error', 'message': 'Brak uprawnień'}, status=403)
        
        # Parse dates
        period_start = request.POST.get('period_start')
        period_end = request.POST.get('period_end')
        
        if not period_start or not period_end:
            return JsonResponse({'status': 'error', 'message': 'Wymagane są daty początkowa i końcowa'}, status=400)
        
        try:
            from datetime import datetime as dt
            period_start_date = dt.strptime(period_start, '%Y-%m-%d').date()
            period_end_date = dt.strptime(period_end, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Nieprawidłowy format daty'}, status=400)
        
        logger.info(f"Generating organization report for period: {period_start_date} to {period_end_date}")
        
        # Build query for tickets in date range
        tickets_query = Ticket.objects.filter(
            created_at__date__gte=period_start_date,
            created_at__date__lte=period_end_date
        ).exclude(
            actual_resolution_time__isnull=True
        )
        
        # Group by organization and calculate average actual resolution time
        org_stats = []
        
        if role == 'admin':
            organizations = Organization.objects.all()
        else:  # superagent
            organizations = user.profile.organizations.all()
        
        for org in organizations:
            org_tickets = tickets_query.filter(organization=org)
            
            if org_tickets.exists():
                total_time = org_tickets.aggregate(
                    total_actual=Sum('actual_resolution_time')
                )['total_actual']
                
                tickets_count = org_tickets.count()
                
                org_stats.append({
                    'name': org.name,
                    'total_actual_time': float(total_time) if total_time else 0,
                    'tickets_count': tickets_count
                })
        
        # Sort by total time (descending)
        org_stats.sort(key=lambda x: x['total_actual_time'], reverse=True)
        
        # Generate Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Raport firm"
        
        # Styles
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=12)
        bold_font = Font(bold=True)
        
        # Title
        ws['A1'] = 'RAPORT RZECZYWISTYCH CZASÓW OBSŁUGI FIRM'
        ws['A1'].font = header_font
        ws.merge_cells('A1:C1')
        
        # Period info
        ws['A3'] = f'Okres: {period_start_date} - {period_end_date}'
        ws['A3'].font = bold_font
        
        ws['A4'] = f'Data generacji: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'
        
        # Headers
        row = 6
        headers = ['Firma', 'Suma rzeczywistego czasu (godz.)', 'Liczba zgłoszeń']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = bold_font
        
        # Data
        row += 1
        for org_data in org_stats:
            ws[f'A{row}'] = org_data['name']
            ws[f'B{row}'] = f"{org_data['total_actual_time']:.2f}"
            ws[f'C{row}'] = org_data['tickets_count']
            row += 1
        
        # Auto-adjust column widths
        for col_idx, column_cells in enumerate(ws.columns, 1):
            length = max(len(str(cell.value or '')) for cell in column_cells if hasattr(cell, 'value'))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(length + 2, 50)
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"raport_firm_{period_start_date}_{period_end_date}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        logger.info(f"Organization report generated successfully: {filename}")
        
        return response
        
    except Exception as e:
        logger.error(f"Error generating organization report: {e}")
        import traceback
        logger.error(f"Full traceback: {traceback.format_exc()}")
        return JsonResponse({
            'status': 'error',
            'message': f'Błąd generowania raportu: {str(e)}'
        }, status=500)
